productsPattern = ['Kindle 10a. geração com bateria de longa duração - Cor Preta',
                   'Suporte para Notebook, OCTOO, Uptable, UP-BL, Preto',
                   'Escrivaninha Trevalla Kuadra Me150-E10 Industrial 150cm Preto Onix',
                   'Echo Dot (4ª Geração): Smart Speaker com Alexa - Cor Preta',
                   'Smart Lâmpada Inteligente Intelbras EWS 410 com 16 milhões de cores, 10W, Casa Inteligente Wi-Fi, compatível com Alexa'
                   ]


def project(product=productsPattern):
    listPrices, product = scrapping(product)
    excel(product, listPrices)


def scrapping(product=productsPattern):
    from selenium import webdriver

    way = 'https://www.amazon.com.br/s?k='
    quantityProducts = len(product)

    chromeOptions = webdriver.ChromeOptions()
    chromeOptions.add_argument('headless')
    driver = webdriver.Chrome(options=chromeOptions)

    listPrices = []

    for index in range(quantityProducts):
        search = str(way) + str(product[index])
        driver.get(search)

        results = driver.find_element_by_css_selector('.s-widget-spacing-small')
        firstProduct = results.find_element_by_css_selector('.sg-col-inner')

        wholePrice = firstProduct.find_element_by_css_selector('.a-price-whole').text
        fractionaryPrice = firstProduct.find_element_by_css_selector('.a-price-fraction').text

        allPrice = float(str(wholePrice) + '.' + str(fractionaryPrice))
        listPrices.append(allPrice)

    return listPrices, product


def excel(product, listPrices):
    from pathlib import Path
    from openpyxl import Workbook
    import pandas as pd

    try:
        Path(r'C:/Users/Henrique/Desktop/Folder').mkdir()  # Make sure there are a folder
    except:
        pass

    folder = Path(r'C:/Users/Henrique/Desktop/Folder').iterdir()

    exist = False
    for file in folder:
        if str(file) == r'C:\Users\Henrique\Desktop\Folder\Amazon´s Produts.xlsx':
            exist = True
            break

    if exist == False:
        wb = Workbook()
        ws = wb.active

        for index, value in enumerate(product):
            ws.cell(row=1, column=index + 1, value=value)
            ws.cell(row=2, column=index + 1, value=listPrices[index])

        wb.save(r'C:\Users\Henrique\Desktop\Folder\Amazon´s Produts.xlsx')

    else:
        excelDF = pd.read_excel(r'C:\Users\Henrique\Desktop\Folder\Amazon´s Produts.xlsx')
        newIndex = excelDF.index.max() + 1

        for index, value in enumerate(listPrices):
            excelDF.loc[newIndex, product[index]] = value

        excelDF.to_excel(r'C:\Users\Henrique\Desktop\Folder\Amazon´s Produts.xlsx', index=False)


project()